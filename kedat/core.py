"""
Defines the core functions and classes for kedat.
"""
import os
from openpyxl import load_workbook, Workbook



class Storage(dict):
    """Represents a dictionary object whose elements can be accessed and set 
    using the dot object notation. Thus in addition to `foo['bar']`, `foo.bar`
    can equally be used.
    """
    
    def __getattr__(self, key):
        return self.__getitem__(key)
    
    def __getitem__(self, key):
        return dict.get(self, key, None)
    
    def __getstate__(self):
        return dict(self)
    
    def __setattr__(self, key, value):
        self[key] = value
    
    def __setstate__(self, value):
        dict.__init__(self, value)
    
    def __repr__(self):
        return "<Storage %s>" % dict.__repr__(self)


class XlSheet:
    """Represents a light wrapper around openpyxl's Worksheet object. Provides
    convenient ways of iterating over rows which are presented as tuples.
    """
    max_row_check = 10
    
    def __init__(self, source, sheet_name, row_offset=0, col_offset=0):
        wb = source if type(source) is Workbook else None
        if not wb and type(source) is str:
            if not os.path.isfile(source):
                raise FileNotFoundError(source)
            
            wb = load_workbook(source, read_only=False)
        
        if not wb:
            raise ValueError(
                "Expected types for source: str, Workbook. Type provided: %s" %
                (type(source),)
            )
        
        wb_sheet_names = wb.get_sheet_names()
        if not sheet_name in wb_sheet_names:
            raise ValueError(
                "Sheet by the name '%s' not found in '%s'." %
                (sheet_name, wb_sheet_names)
            )
        
        self.worksheet = wb[sheet_name]
        self.sheet_name = sheet_name
        self.__col_offset = col_offset
        self.__row_offset = row_offset
        self.__generator = None
        self.__current = None
    
    @property
    def current(self):
        return self.__current
    
    @property
    def max_column(self):
        return self.worksheet.max_column
    
    @property
    def max_row(self):
        return self.worksheet.max_row
    
    @property
    def col_offset(self):
        return self.__col_offset
    
    @col_offset.setter
    def col_offset(self, value):
        self.__col_offset = value
        self.__generator = None
    
    @property
    def row_offset(self):
        return self.__row_offset
    
    @row_offset.setter
    def row_offset(self, value):
        self.__row_offset = value
        self.__generator = None
    
    def iter_rows(self, row_offset=0):
        self.row_offset = row_offset
        return self.__get_generator()
    
    def next(self):
        return self.__get_generator().send(None)
    
    def reset(self):
        self.__generator = None
    
    def __iter__(self):
        return self.__get_generator()
    
    def __get_generator(self):
        def make_generator():
            for i in range(self.row_offset + 1, self.max_row + 1):
                row = []
                for j in range(self.col_offset + 1, self.max_column + 1):
                    value = self.worksheet.cell(row=i, column=j).value
                    row.append(value)
                
                self.__current = row = tuple(row)
                yield row
                
        if self.__generator is None:
            self.__generator = make_generator()
        return self.__generator
    
    @staticmethod
    def find_headers(xlsheet, sample_headers, row_offset=0):
        norm_hdrs = [h.lower() for h in sample_headers]
        hdr_count = len(norm_hdrs)
        
        idx = row_offset
        xlsheet.row_offset = row_offset
        for row in xlsheet:
            idx += 1
            
            norm_row = [str(c or '').lower() for c in row]
            if norm_row[:hdr_count] == norm_hdrs:
                return idx
            
            if (idx - row_offset) == XlSheet.max_row_check:
                return -1
            
            